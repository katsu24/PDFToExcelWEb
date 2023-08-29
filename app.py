import io
import os
from flask import Flask, render_template, request, send_file
import pandas as pd
import tabula
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import fitz
from docx import Document
import logging
import subprocess

app = Flask(__name__)

# ログの設定
log_file = os.path.join(os.path.dirname(__file__), 'app.log')
handler = logging.FileHandler(log_file)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route("/pdf_to")
def index():
    return render_template("upload.html")

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    pdf_file = request.files["pdf_file"]
    if pdf_file:
        dfs = tabula.read_pdf(pdf_file, lattice=True, pages='all')
        combined_df = pd.concat(dfs, ignore_index=True)

        excel_io = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active

        for row in dataframe_to_rows(combined_df, index=False, header=True):
            sheet.append(row)

        # セル内の改行に対して折り返し設定を行う
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # セル幅を調整
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # 余白を考慮して調整
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(excel_io)
        excel_io.seek(0)
    
        return send_file(
            excel_io,
            as_attachment=True,
            download_name='result.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
@app.route("/upload_word", methods=["POST"])
def upload_word():
    pdf_file = request.files["pdf_file"]
    if pdf_file:
        # PDFファイルを読み込み、テキストを抽出
        pdf_data = pdf_file.read()
        pdf_document = fitz.open(stream=pdf_data, filetype="pdf")
        text = ""
        for page in pdf_document:
            text += page.get_text("text")

        document = Document()
        document.add_paragraph(text)
        
        word_io = io.BytesIO()
        document.save(word_io)
        word_io.seek(0)

        return send_file(
            word_io,
            as_attachment=True,
            download_name='result.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

if __name__ == "__main__":
    app.run('0.0.0.0')